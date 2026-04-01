"""
results_exporter.py
===================
TITLE
Formatted workbook exporter for pipeline outputs.

PURPOSE
This module turns raw CSV and JSON artefacts into a presentation-ready Excel
workbook without changing the scientific content of the pipeline outputs.

WHAT IT DOES
- Loads results, uncertainty, calibration, and dataset artefacts.
- Writes a multi-sheet Excel workbook with formatting and summaries.
- Preserves source metrics for downstream review.

HOW IT WORKS
1. Load result tables from the pipeline output directory.
2. Build workbook sheets for summary, seeds, kinase metrics, and calibration.
3. Apply formatting and save a single workbook file.

INPUT CONTRACT
- Existing result CSV/JSON files under a results directory.
- Dataset parquet for dataset statistics.

OUTPUT CONTRACT
- One Excel workbook containing the available reporting sheets.

DEPENDENCIES
- pandas, numpy, openpyxl

CRITICAL ASSUMPTIONS
- Upstream result files are already trustworthy and version-compatible.

FAILURE MODES
- Missing source files
- Workbook dependency issues
- Empty optional result tables

SAFETY CHECKS IMPLEMENTED
- Optional-sheet handling when some files are absent
- File-type aware loading helpers
- Non-destructive export behaviour

HOW TO RUN
- `python results_exporter.py --results-dir ./pipeline_outputs/results --output pipeline_results.xlsx`

HOW IT CONNECTS TO PIPELINE
It is the final reporting layer fed by evaluation and experiment artefacts.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

log = logging.getLogger("results_exporter")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Colour palette (openpyxl hex, no #)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"   # white
C_BEST_BG     = "E2EFDA"   # light green â€” best value in column
C_WARN_BG     = "FCE4D6"   # light orange â€” worst value
C_SUBHEAD_BG  = "D6E4F0"   # light blue â€” section headers
C_ALT_ROW     = "F2F2F2"   # light grey â€” alternating rows
C_BORDER      = "8EA9C1"   # medium blue â€” thin border colour

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# openpyxl helpers
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _get_xl():
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
        return openpyxl, PatternFill, Font, Alignment, Border, Side, get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export.\n"
            "Install with: pip install openpyxl"
        )


def _header_fill(color: str = C_HEADER_BG):
    _, PatternFill, *_ = _get_xl()
    return PatternFill("solid", fgColor=color)


def _header_font(bold: bool = True, color: str = C_HEADER_FG, size: int = 11):
    _, _, Font, *_ = _get_xl()
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _center():
    _, _, _, Alignment, *_ = _get_xl()
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    openpyxl, _, _, _, Border, Side, _ = _get_xl()
    thin = Side(style="thin", color=C_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(ws, row: int, n_cols: int, bg: str = C_HEADER_BG) -> None:
    """Apply header styling to an entire row."""
    openpyxl, *_ = _get_xl()
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = _header_fill(bg)
        cell.font   = _header_font(color=C_HEADER_FG if bg == C_HEADER_BG else "1F3864")
        cell.alignment = _center()
        cell.border = _thin_border()


def _style_data_row(ws, row: int, n_cols: int, alt: bool = False) -> None:
    """Apply alternating-row background to a data row."""
    openpyxl, PatternFill, Font, Alignment, Border, Side, _ = _get_xl()
    bg = C_ALT_ROW if alt else "FFFFFF"
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()


def _autofit_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Set column widths based on content length."""
    openpyxl, _, _, _, _, _, get_column_letter = _get_xl()
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(
            max(length + 2, min_width), max_width
        )


def _freeze_top_row(ws) -> None:
    ws.freeze_panes = ws["A2"]


def _highlight_best_worst(ws, data_start_row: int, col: int,
                           n_rows: int, higher_is_better: bool = True) -> None:
    """Highlight the best (green) and worst (orange) cell in a numeric column."""
    openpyxl, PatternFill, *_ = _get_xl()
    values = []
    for r in range(data_start_row, data_start_row + n_rows):
        v = ws.cell(row=r, column=col).value
        try:
            values.append((float(v), r))
        except (TypeError, ValueError):
            pass
    if not values:
        return
    best_row = max(values, key=lambda x: x[0])[1] if higher_is_better \
        else min(values, key=lambda x: x[0])[1]
    worst_row = min(values, key=lambda x: x[0])[1] if higher_is_better \
        else max(values, key=lambda x: x[0])[1]
    ws.cell(row=best_row,  column=col).fill = PatternFill("solid", fgColor=C_BEST_BG)
    ws.cell(row=worst_row, column=col).fill = PatternFill("solid", fgColor=C_WARN_BG)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Sheet builders
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _write_summary_sheet(wb, results_df: pd.DataFrame) -> None:
    """Sheet 1: Summary â€” one row per config, meanÂ±std metrics."""
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Kinaseâ€“Ligand Prediction â€” Model Comparison Summary"
    ws["A1"].font = _header_font(size=13, color="1F3864")
    ws.merge_cells("A1:L1")
    ws["A1"].alignment = _get_xl()[3](horizontal="center")

    # Column definitions
    cols = [
        ("Rank",           "rank"),
        ("Config ID",      "config_id"),
        ("Spearman Ï",     "spearman_mean"),
        ("Â± std",          "spearman_std"),
        ("RMSE",           "rmse_mean"),
        ("Â± std",          "rmse_std"),
        ("EF1% (proxy?)",  "ef1pct_mean"),
        ("Â± std",          "ef1pct_std"),
        ("Calibration Ï",  "calibration_mean"),
        ("Î” vs Baseline",  "delta_vs_baseline"),
        ("Meaningful?",    "meaningful_gain"),
        ("N seeds",        "n_seeds"),
    ]

    header_row = 3
    for c_idx, (label, _) in enumerate(cols, 1):
        ws.cell(row=header_row, column=c_idx, value=label)
    _style_header_row(ws, header_row, len(cols))

    # Sort by Spearman descending
    if "spearman_mean" in results_df.columns:
        df = results_df.sort_values("spearman_mean", ascending=False).reset_index(drop=True)
    else:
        df = results_df.reset_index(drop=True)

    for r_idx, (_, row) in enumerate(df.iterrows(), 1):
        excel_row = header_row + r_idx
        values = [r_idx] + [row.get(col_key, "") for _, col_key in cols[1:]]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=c_idx)
            if isinstance(val, float) and not np.isnan(val):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
            elif isinstance(val, bool):
                cell.value = "âœ“" if val else "â€“"
            else:
                cell.value = val
        _style_data_row(ws, excel_row, len(cols), alt=(r_idx % 2 == 0))

    # Highlight best/worst for Spearman and RMSE
    n_data = len(df)
    spearman_col = [l for l, k in cols].index("Spearman Ï") + 1
    rmse_col     = [l for l, k in cols].index("RMSE") + 1
    _highlight_best_worst(ws, header_row + 1, spearman_col, n_data, higher_is_better=True)
    _highlight_best_worst(ws, header_row + 1, rmse_col,     n_data, higher_is_better=False)

    # Legend
    legend_row = header_row + n_data + 3
    ws.cell(row=legend_row,     column=1, value="Legend:").font = _header_font(color="1F3864", size=10)
    ws.cell(row=legend_row + 1, column=1, value="Green = best value in column")
    ws.cell(row=legend_row + 2, column=1, value="Orange = worst value in column")
    ws.cell(row=legend_row + 3, column=1,
            value="EF1% marked as proxy if <5% inactive background â€” not a valid VS metric.")
    ws.cell(row=legend_row + 4, column=1,
            value="Meaningful? = True only if Î” > seed-std (decision rule applied).")

    _autofit_columns(ws)
    _freeze_top_row(ws)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[header_row].height = 35


def _write_per_seed_sheet(wb, per_seed_df: pd.DataFrame) -> None:
    """Sheet 2: Per_Seed â€” raw metric per (config, seed)."""
    ws = wb.create_sheet("Per_Seed")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Per-Seed Raw Metrics"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:I1")

    cols = [
        "config_id", "seed", "spearman", "rmse",
        "ef1pct", "calibration", "n_test", "val_loss",
    ]
    for c, name in enumerate(cols, 1):
        ws.cell(row=3, column=c, value=name)
    _style_header_row(ws, 3, len(cols))

    for r, (_, row) in enumerate(per_seed_df.iterrows(), 1):
        excel_row = 3 + r
        for c, col in enumerate(cols, 1):
            val = row.get(col, "")
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(val, float) and not (val != val):  # not NaN
                cell.value = round(val, 5)
                cell.number_format = "0.00000"
            else:
                cell.value = val
        _style_data_row(ws, excel_row, len(cols), alt=(r % 2 == 0))

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_per_kinase_sheet(wb, per_kinase_df: pd.DataFrame) -> None:
    """Sheet 3: Per_Kinase â€” per-UniProt evaluation."""
    ws = wb.create_sheet("Per_Kinase")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Per-Kinase Evaluation (configs with â‰¥ 3 test samples per kinase)"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:G1")

    if per_kinase_df.empty:
        ws["A3"] = "No per-kinase data available."
        return

    cols = ["config_id", "uniprot_id", "n_samples", "spearman", "rmse", "calibration"]
    available_cols = [c for c in cols if c in per_kinase_df.columns]

    for c, name in enumerate(available_cols, 1):
        ws.cell(row=3, column=c, value=name)
    _style_header_row(ws, 3, len(available_cols))

    for r, (_, row) in enumerate(per_kinase_df.iterrows(), 1):
        excel_row = 3 + r
        for c, col in enumerate(available_cols, 1):
            val = row.get(col, "")
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(val, float) and not (val != val):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
            else:
                cell.value = val
        _style_data_row(ws, excel_row, len(available_cols), alt=(r % 2 == 0))

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_uncertainty_sheet(wb, unc_df: pd.DataFrame) -> None:
    """Sheet 4: Uncertainty â€” ensemble predictions with CIs on test set."""
    ws = wb.create_sheet("Uncertainty")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Ensemble Uncertainty Predictions (Test Set)"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:J1")

    cols = [
        "inchikey", "uniprot_id", "target", "pred_mean", "pred_std",
        "aleatoric_std", "epistemic_std", "lower_95", "upper_95",
    ]
    available = [c for c in cols if c in unc_df.columns]

    for c, name in enumerate(available, 1):
        ws.cell(row=3, column=c, value=name)
    _style_header_row(ws, 3, len(available))

    # Sort by pred_mean descending (most active first â€” useful for hit-list)
    if "pred_mean" in unc_df.columns:
        unc_df = unc_df.sort_values("pred_mean", ascending=False).reset_index(drop=True)

    for r, (_, row) in enumerate(unc_df.iterrows(), 1):
        excel_row = 3 + r
        for c, col in enumerate(available, 1):
            val = row.get(col, "")
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(val, float) and not (val != val):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
            else:
                cell.value = val
        _style_data_row(ws, excel_row, len(available), alt=(r % 2 == 0))

    # Conditional format: highlight rows where target outside CI
    note_row = len(unc_df) + 5
    ws.cell(row=note_row, column=1,
            value="Rows sorted by predicted pIC50 (descending) â€” highest-confidence hits first.")

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_calibration_sheet(wb, cal_metrics: dict, rel_df: Optional[pd.DataFrame] = None) -> None:
    """Sheet 5: Calibration â€” coverage table and ECE."""
    ws = wb.create_sheet("Calibration")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Uncertainty Calibration Metrics"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:D1")

    # Scalar metrics table
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    ws["C3"] = "Interpretation"
    _style_header_row(ws, 3, 3)

    interpretations = {
        "spearman_err_sigma": "Higher is better (>0.3 = reasonable calibration)",
        "coverage_95":        "Ideal = 0.95; <0.95 = overconfident",
        "ece":                "Lower is better (0 = perfect calibration)",
        "sharpness":          "Lower = sharper predictions (narrower CIs)",
        "mean_nll":           "Lower is better (Gaussian NLL on test set)",
        "mean_abs_error":     "Mean absolute prediction error (pIC50 units)",
    }
    scalar_keys = [k for k, v in cal_metrics.items() if isinstance(v, float)]
    for r, key in enumerate(scalar_keys, 1):
        val = cal_metrics[key]
        excel_row = 3 + r
        ws.cell(row=excel_row, column=1, value=key)
        ws.cell(row=excel_row, column=2, value=round(val, 5) if not np.isnan(val) else "N/A")
        ws.cell(row=excel_row, column=2).number_format = "0.00000"
        ws.cell(row=excel_row, column=3, value=interpretations.get(key, ""))
        _style_data_row(ws, excel_row, 3, alt=(r % 2 == 0))

    # Reliability diagram table
    if rel_df is not None and not rel_df.empty:
        offset = len(scalar_keys) + 6
        ws.cell(row=offset, column=1, value="Reliability Diagram Data")
        ws.cell(row=offset, column=1).font = _header_font(size=11, color="1F3864")

        ws.cell(row=offset + 1, column=1, value="Expected Coverage")
        ws.cell(row=offset + 1, column=2, value="Observed Coverage")
        ws.cell(row=offset + 1, column=3, value="Count")
        _style_header_row(ws, offset + 1, 3)

        for r, (_, row) in enumerate(rel_df.iterrows(), 1):
            er = offset + 1 + r
            ws.cell(row=er, column=1, value=round(row["expected_coverage"], 2))
            ws.cell(row=er, column=2, value=round(row["observed_coverage"], 4))
            ws.cell(row=er, column=3, value=int(row.get("count", 0)))
            ws.cell(row=er, column=1).number_format = "0.00"
            ws.cell(row=er, column=2).number_format = "0.0000"
            _style_data_row(ws, er, 3, alt=(r % 2 == 0))

    _autofit_columns(ws)


def _write_pairwise_sheet(wb, pairwise_df: pd.DataFrame) -> None:
    """Sheet 6: Pairwise_Stats â€” Wilcoxon significance between configs."""
    ws = wb.create_sheet("Pairwise_Stats")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Pairwise Wilcoxon Signed-Rank Tests (Spearman Ï across seeds)"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:G1")

    ws["A2"] = ("âš  With only 3 seeds, statistical power is limited. "
                "p-values are directional, not definitive.")
    ws["A2"].font = _header_font(bold=False, size=10, color="C00000")
    ws.merge_cells("A2:G2")

    if pairwise_df.empty:
        ws["A4"] = "No pairwise tests computed."
        return

    cols = ["config_a", "config_b", "delta_mean", "p_value", "significant", "better"]
    available = [c for c in cols if c in pairwise_df.columns]

    for c, name in enumerate(available, 1):
        ws.cell(row=4, column=c, value=name)
    _style_header_row(ws, 4, len(available))

    sig_col = available.index("significant") + 1 if "significant" in available else None

    for r, (_, row) in enumerate(pairwise_df.iterrows(), 1):
        excel_row = 4 + r
        for c, col in enumerate(available, 1):
            val = row.get(col, "")
            cell = ws.cell(row=excel_row, column=c)
            if col == "significant":
                cell.value = "âœ“" if val else "â€“"
            elif isinstance(val, float) and not (val != val):
                cell.value = round(val, 5)
                cell.number_format = "0.00000"
            else:
                cell.value = val
        _style_data_row(ws, excel_row, len(available), alt=(r % 2 == 0))
        # Highlight significant rows
        if sig_col and row.get("significant", False):
            for c in range(1, len(available) + 1):
                from openpyxl.styles import PatternFill
                ws.cell(row=excel_row, column=c).fill = PatternFill(
                    "solid", fgColor=C_BEST_BG
                )

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_dataset_stats_sheet(wb, dataset_path: str) -> None:
    """Sheet 7: Dataset_Stats â€” pIC50 distribution, kinase class breakdown."""
    ws = wb.create_sheet("Dataset_Stats")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Dataset Statistics"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:E1")

    try:
        df = pd.read_parquet(dataset_path)
    except Exception as e:
        ws["A3"] = f"Could not load dataset: {e}"
        return

    # Global stats
    stats = {
        "Total rows":                  len(df),
        "Unique molecules (InChIKey)": df["inchikey"].nunique(),
        "Unique kinase targets":       df["uniprot_id"].nunique(),
        "pIC50 mean":                  round(df["pIC50"].mean(), 3),
        "pIC50 std":                   round(df["pIC50"].std(), 3),
        "pIC50 min":                   round(df["pIC50"].min(), 3),
        "pIC50 max":                   round(df["pIC50"].max(), 3),
        "pIC50 median":                round(df["pIC50"].median(), 3),
        "Active (â‰¥6.0)":               int((df["pIC50"] >= 6.0).sum()),
        "Inactive (<6.0)":             int((df["pIC50"] < 6.0).sum()),
        "Active %":                    f"{100*(df['pIC50']>=6.0).mean():.1f}%",
        "Singleton measurements":      int((df["n_measurements"] == 1).sum()),
        "Multi-measurement":           int((df["n_measurements"] > 1).sum()),
    }

    ws.cell(row=3, column=1, value="Metric")
    ws.cell(row=3, column=2, value="Value")
    _style_header_row(ws, 3, 2)

    for r, (k, v) in enumerate(stats.items(), 1):
        ws.cell(row=3 + r, column=1, value=k)
        ws.cell(row=3 + r, column=2, value=v)
        _style_data_row(ws, 3 + r, 2, alt=(r % 2 == 0))

    # Kinase class breakdown
    if "kinase_class" in df.columns:
        offset = len(stats) + 6
        ws.cell(row=offset, column=1, value="Kinase Class Breakdown")
        ws.cell(row=offset, column=1).font = _header_font(size=11, color="1F3864")
        ws.merge_cells(f"A{offset}:C{offset}")
        ws.cell(row=offset + 1, column=1, value="Class")
        ws.cell(row=offset + 1, column=2, value="Unique Targets")
        ws.cell(row=offset + 1, column=3, value="Total Records")
        _style_header_row(ws, offset + 1, 3)
        class_stats = df.groupby("kinase_class").agg(
            unique_targets=("uniprot_id", "nunique"),
            total_records=("pIC50", "count"),
        ).reset_index()
        for r, row in enumerate(class_stats.itertuples(index=False), 1):
            er = offset + 1 + r
            ws.cell(row=er, column=1, value=row.kinase_class)
            ws.cell(row=er, column=2, value=row.unique_targets)
            ws.cell(row=er, column=3, value=row.total_records)
            _style_data_row(ws, er, 3, alt=(r % 2 == 0))

    # pIC50 histogram bins
    if "pIC50" in df.columns:
        offset2 = (len(stats) + 6 + len(class_stats) + 4
                   if "kinase_class" in df.columns else len(stats) + 6)
        ws.cell(row=offset2, column=1, value="pIC50 Histogram (1-unit bins)")
        ws.cell(row=offset2, column=1).font = _header_font(size=11, color="1F3864")
        ws.merge_cells(f"A{offset2}:C{offset2}")
        ws.cell(row=offset2 + 1, column=1, value="pIC50 Range")
        ws.cell(row=offset2 + 1, column=2, value="Count")
        ws.cell(row=offset2 + 1, column=3, value="% of Total")
        _style_header_row(ws, offset2 + 1, 3)
        bins = np.arange(2, 16)
        counts, _ = np.histogram(df["pIC50"].dropna(), bins=bins)
        for r, (lo, hi, cnt) in enumerate(zip(bins[:-1], bins[1:], counts), 1):
            er = offset2 + 1 + r
            ws.cell(row=er, column=1, value=f"[{lo:.0f}, {hi:.0f})")
            ws.cell(row=er, column=2, value=int(cnt))
            ws.cell(row=er, column=3, value=f"{100*cnt/len(df):.1f}%")
            _style_data_row(ws, er, 3, alt=(r % 2 == 0))

    _autofit_columns(ws)


def _write_feature_inventory_sheet(wb) -> None:
    """Sheet 8: Feature_Inventory â€” documents all feature dimensions."""
    ws = wb.create_sheet("Feature_Inventory")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Feature Engineering Inventory"
    ws["A1"].font = _header_font(size=12, color="1F3864")
    ws.merge_cells("A1:D1")

    sections = [
        ("LIGAND GRAPH FEATURES (module2)", [
            ("Atom features â€” element (14-way one-hot)",   "43-dim total", "per atom"),
            ("Atom features â€” degree (8-way one-hot)",     "â†‘ included",  "per atom"),
            ("Atom features â€” formal charge (8-way)",      "â†‘ included",  "per atom"),
            ("Atom features â€” hybridisation (6-way)",      "â†‘ included",  "per atom"),
            ("Atom features â€” chirality (4-way one-hot)",  "â†‘ included",  "per atom"),
            ("Atom features â€” is_aromatic (binary)",       "â†‘ included",  "per atom"),
            ("Atom features â€” is_in_ring (binary)",        "â†‘ included",  "per atom"),
            ("Atom features â€” num_Hs (0â€“4, norm.)",        "â†‘ included",  "per atom"),
            ("Bond features â€” bond type (5-way)",          "12-dim total","per bond"),
            ("Bond features â€” conjugated (binary)",        "â†‘ included",  "per bond"),
            ("Bond features â€” in_ring (binary)",           "â†‘ included",  "per bond"),
            ("Bond features â€” stereo E/Z/any (5-way)",     "â†‘ included",  "per bond"),
            ("Edges are bidirectional (iâ†’j AND jâ†’i)",      "â€”",           "graph"),
        ]),
        ("LIGAND PHYSICOCHEMICAL FEATURES (module2)", [
            ("MolWt, ExactMolWt, MolLogP, TPSA",          "4 scalars",  "per mol"),
            ("NumHDonors, NumHAcceptors, NumRotBonds",     "3 scalars",  "per mol"),
            ("NumAromaticRings, NumAliphaticRings",        "2 scalars",  "per mol"),
            ("NumHeavyAtoms, NumSatRings, FractionCSP3",   "3 scalars",  "per mol"),
            ("LabuteASA, BalabanJ, BertzCT",               "3 scalars",  "per mol"),
            ("Chi0n, Chi1n, Kappa1, Kappa2",               "4 scalars",  "per mol"),
            ("RingCount, NumRadicalElec, MaxPartialCharge","3 scalars",  "per mol"),
            ("Total physchem dim",                         "22",         "per mol"),
        ]),
        ("MORGAN FINGERPRINT (module2 â€” ECFP4)", [
            ("Morgan fingerprint, radius=2, nBits=1024",   "1024-bit",   "per mol"),
            ("Stored as data.morgan_fp on PyG Data object","â€”",          "per mol"),
            ("Used by MLPLigandEncoder (protein_only ablation)", "â€”",   "per mol"),
        ]),
        ("PROTEIN FEATURES (module3)", [
            ("ESM-2 (esm2_t33_650M_UR50D) full-sequence", "â€”",           "per target"),
            ("Slice at KLIFS 85 pocket indices",          "85 Ã— 1280",  "per target"),
            ("AlphaFold CÎ± coordinates (centred)",        "85 Ã— 3",     "per target"),
            ("AlphaFold pLDDT scores",                    "85",         "per target"),
            ("Confidence = pLDDT / 100",                  "85",         "per target"),
            ("Pocket mask (True = valid residue)",        "85 bool",    "per target"),
        ]),
    ]

    current_row = 3
    for section_title, rows in sections:
        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = _header_font(size=11, color="1F3864")
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws.cell(row=current_row, column=1).fill = _header_fill(C_SUBHEAD_BG)
        current_row += 1

        ws.cell(row=current_row, column=1, value="Feature")
        ws.cell(row=current_row, column=2, value="Dimension")
        ws.cell(row=current_row, column=3, value="Scope")
        _style_header_row(ws, current_row, 3)
        current_row += 1

        for r, (feat, dim, scope) in enumerate(rows):
            ws.cell(row=current_row, column=1, value=feat)
            ws.cell(row=current_row, column=2, value=dim)
            ws.cell(row=current_row, column=3, value=scope)
            _style_data_row(ws, current_row, 3, alt=(r % 2 == 0))
            current_row += 1

        current_row += 1  # blank row between sections

    _autofit_columns(ws)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Main export function
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def export_results_workbook(
    results_dir:    str = "./pipeline_outputs/results",
    dataset_path:   str = "./pipeline_outputs/dataset_clean.parquet",
    output_path:    str = "./pipeline_outputs/pipeline_results.xlsx",
) -> str:
    """
    Build and save the full results Excel workbook.

    Parameters
    ----------
    results_dir  : directory containing results CSV files from module9
    dataset_path : path to dataset_clean.parquet (for Dataset_Stats sheet)
    output_path  : output .xlsx file path

    Returns
    -------
    str : absolute path to the saved workbook
    """
    openpyxl, *_ = _get_xl()
    wb = openpyxl.Workbook()

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    results_dir = Path(results_dir)

    # â”€â”€ Load all result files â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    def _try_load(fname: str, default=None):
        path = results_dir / fname
        if path.exists():
            try:
                if fname.endswith(".csv"):
                    return pd.read_csv(path)
                elif fname.endswith(".json"):
                    import json
                    with open(path) as f:
                        return json.load(f)
            except Exception as e:
                log.warning("Could not load %s: %s", path, e)
        return default

    results_df   = _try_load("results.csv",                  pd.DataFrame())
    per_seed_df  = _try_load("per_seed_results.csv",         pd.DataFrame())
    per_kinase_df= _try_load("per_kinase_results.csv",       pd.DataFrame())
    unc_df       = _try_load("uncertainty_predictions.csv",  pd.DataFrame())
    cal_metrics  = _try_load("calibration_metrics.json",     {})
    rel_df       = _try_load("reliability_diagram.csv",      None)
    pairwise_df  = _try_load("pairwise_significance.csv",    pd.DataFrame())

    # â”€â”€ Build sheets â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    log.info("Building Summary sheet â€¦")
    _write_summary_sheet(wb, results_df)

    log.info("Building Per_Seed sheet â€¦")
    _write_per_seed_sheet(wb, per_seed_df)

    log.info("Building Per_Kinase sheet â€¦")
    _write_per_kinase_sheet(wb, per_kinase_df)

    log.info("Building Uncertainty sheet â€¦")
    if not unc_df.empty:
        _write_uncertainty_sheet(wb, unc_df)

    log.info("Building Calibration sheet â€¦")
    _write_calibration_sheet(wb, cal_metrics, rel_df)

    log.info("Building Pairwise_Stats sheet â€¦")
    _write_pairwise_sheet(wb, pairwise_df)

    log.info("Building Dataset_Stats sheet â€¦")
    _write_dataset_stats_sheet(wb, dataset_path)

    log.info("Building Feature_Inventory sheet â€¦")
    _write_feature_inventory_sheet(wb)

    # â”€â”€ Save â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    output_path = str(Path(output_path).resolve())
    wb.save(output_path)
    log.info("Workbook saved â†’ %s  (%d sheets)", output_path, len(wb.sheetnames))
    return output_path


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Entry point
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Export all pipeline results to a formatted Excel workbook."
    )
    parser.add_argument("--results-dir", default="./pipeline_outputs/results",
                        help="Directory containing CSV/JSON result files")
    parser.add_argument("--dataset",     default="./pipeline_outputs/dataset_clean.parquet",
                        help="Path to dataset_clean.parquet")
    parser.add_argument("--output",      default="./pipeline_outputs/pipeline_results.xlsx",
                        help="Output Excel file path")
    args = parser.parse_args()

    path = export_results_workbook(
        results_dir  = args.results_dir,
        dataset_path = args.dataset,
        output_path  = args.output,
    )
    print(f"\nResults workbook â†’ {path}")
    print("Sheets: Summary | Per_Seed | Per_Kinase | Uncertainty | "
          "Calibration | Pairwise_Stats | Dataset_Stats | Feature_Inventory")

